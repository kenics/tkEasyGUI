import os
from openpyxl import load_workbook
from datetime import datetime

folder_path = r"\\File2024\PJ_信越FA\01.現行システム\01.プログラム仕様書"
keyword = "密度"

# exclude_keyword_list = ["更新日時","新規"]
exclude_keyword_list = ["(","断面積","管理"]

current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
output_file_name = f"search_results_{keyword}_{current_time}.csv"
output_dir = "results"
keyword_list_file = f"keyword_list_file.txt"


def check_exclude(value):
	return any(exclude in value for exclude in exclude_keyword_list)


excel_files = [f for f in os.listdir(folder_path) if f.startswith("データベース定義書") and f.endswith(('.xlsx'))]

results = []

for file in excel_files:
	file_path = os.path.join(folder_path,file)
	try:
		workbook = load_workbook(file_path,data_only=True)
		for sheet_name in workbook.sheetnames:
			sheet = workbook[sheet_name]
			for row in sheet.iter_rows():
				for cell in row:
					if cell.value and keyword in str(cell.value) and not check_exclude(str(cell.value)):
						results.append({
							"file":file,
							"sheet":sheet_name,
							"cell":cell.coordinate,
							"value":cell.value
						})
	except Exception as e:
		print(f"Error{file}:{e}")
output_file = os.path.join(output_dir,output_file_name)
with open(output_file,"w",encoding="cp932") as f:
	if results:
		for result in results:
			f.write(f"File:{result['file']},Sheet:{result['sheet']},Cell:{result['cell']},Value:{result['value']}\n")
		print(f"{output_file_name}")
	else:
		print(f"not found Keyword:{keyword}")

with open(keyword_list_file,"a",encoding="utf-8") as f:
	if results:
		f.write(f"{keyword} 除外:{exclude_keyword_list}\n")
	else:
		f.write(f"not found Keyword:{keyword} 除外:{exclude_keyword_list}\n")

